"""Tests for :mod:`src.ingest`.

Every spreadsheet here is written by :func:`workbook` into a temporary
directory, so the suite needs none of the 850 MB of source data and stays
fast. The rows are tiny; what is being tested is the handling around them —
which sheet gets read, which rows get dropped, and what happens when the
Parquet cache is missing, stale or damaged.

The cache paths matter more than they look. Reading the nine real files takes
about 15 minutes and converting each one 45-130 seconds, so an interruption
partway through is likely rather than theoretical, and a cache that silently
returns the wrong data would change every number the project publishes with
nothing to indicate it.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src import ingest

HEADER = ["CASE_NUMBER", "CASE_STATUS", "DECISION_DATE", "EMPLOYER_NAME"]


def workbook(
    path: Path,
    rows: list[list[object]],
    header: list[str] | None = None,
    sheet_name: str = "Sheet1",
    padding: int = 0,
) -> Path:
    """Write one .xlsx, optionally followed by blank padding rows."""
    book = Workbook()
    sheet = book.active
    sheet.title = sheet_name
    sheet.append(header if header is not None else HEADER)
    for row in rows:
        sheet.append(row)
    for _ in range(padding):
        sheet.append([None] * len(header if header is not None else HEADER))
    book.save(path)
    return path


@pytest.fixture
def source(tmp_path):
    """A raw directory with one two-row spreadsheet, and an interim directory."""
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    workbook(
        raw / "LCA_Disclosure_Data_FY2025_Q1.xlsx",
        [
            ["I-200-001", "Certified", "2025-01-15", "ACME"],
            ["I-200-002", "Denied", "2025-01-16", "BETA"],
        ],
    )
    return raw, interim


# --------------------------------------------------------------------------
# Finding the source files
# --------------------------------------------------------------------------


def test_source_files_are_found_by_glob_not_by_name(tmp_path):
    """DOL published one file as 'Dislclosure'. Matching literal names loses it."""
    for name in (
        "LCA_Disclosure_Data_FY2025_Q1.xlsx",
        "LCA_Dislclosure_Data_FY2026_Q2.xlsx",
    ):
        workbook(tmp_path / name, [["I-200-001", "Certified", "2025-01-15", "ACME"]])
    assert [p.name for p in ingest.source_files(tmp_path)] == [
        "LCA_Disclosure_Data_FY2025_Q1.xlsx",
        "LCA_Dislclosure_Data_FY2026_Q2.xlsx",
    ]


def test_source_files_are_returned_in_a_stable_order(tmp_path):
    names = [f"LCA_Disclosure_Data_FY2025_Q{q}.xlsx" for q in (3, 1, 4, 2)]
    for name in names:
        workbook(tmp_path / name, [["I-200-001", "Certified", "2025-01-15", "ACME"]])
    found = [p.name for p in ingest.source_files(tmp_path)]
    assert found == sorted(names)
    assert found == [p.name for p in ingest.source_files(tmp_path)]


def test_an_empty_raw_directory_says_where_to_get_the_data(tmp_path):
    with pytest.raises(FileNotFoundError, match="README"):
        ingest.source_files(tmp_path)


# --------------------------------------------------------------------------
# Reading a spreadsheet
# --------------------------------------------------------------------------


def test_blank_padding_rows_are_dropped_on_read(tmp_path):
    """73% of the real rows are blank. Counting before dropping them is wrong."""
    path = workbook(
        tmp_path / "f.xlsx",
        [["I-200-001", "Certified", "2025-01-15", "ACME"]],
        padding=500,
    )
    assert load_workbook(path, read_only=True).active.max_row > 100
    assert len(ingest.read_xlsx(path)) == 1


def test_the_sheet_is_chosen_by_position_not_by_name(tmp_path):
    """All nine real files use a different sheet name, and none is 'Sheet1'."""
    path = workbook(
        tmp_path / "f.xlsx",
        [["I-200-001", "Certified", "2025-01-15", "ACME"]],
        sheet_name="LCA_Disclosure_Data_FY2025_Q1",
    )
    assert len(ingest.read_xlsx(path)) == 1


def test_a_spreadsheet_without_a_case_number_column_raises(tmp_path):
    path = workbook(
        tmp_path / "f.xlsx", [["x", "y"]], header=["SOMETHING", "ELSE"]
    )
    with pytest.raises(ValueError, match="no CASE_NUMBER column"):
        ingest.read_xlsx(path)


def test_a_row_is_kept_when_only_its_case_number_is_present(tmp_path):
    """Emptiness is judged on CASE_NUMBER alone, not on the whole row."""
    path = workbook(tmp_path / "f.xlsx", [["I-200-001", None, None, None]])
    assert len(ingest.read_xlsx(path)) == 1


def test_case_number_is_located_by_name_not_by_position(tmp_path):
    """It is the first column in all nine files today, which is why this needs
    a test: nothing else would notice if the lookup were replaced by ``0``.

    The column set already changed once mid-fiscal-year, so the position is
    not a property of the data — it is a coincidence that currently holds.
    """
    path = workbook(
        tmp_path / "f.xlsx",
        [
            ["Certified", "2025-01-15", "I-200-001", "ACME"],
            ["Denied", None, None, None],  # padding: no case number
        ],
        header=["CASE_STATUS", "DECISION_DATE", "CASE_NUMBER", "EMPLOYER_NAME"],
    )
    frame = ingest.read_xlsx(path)
    assert len(frame) == 1
    assert frame["CASE_NUMBER"].iloc[0] == "I-200-001"


def test_the_workbook_is_opened_in_read_only_mode(tmp_path, monkeypatch):
    """Without it openpyxl loads a whole 140 MB spreadsheet into memory.

    A proxy for the real property rather than a measurement of it: the
    fixtures here are far too small for the difference to show up, so the
    argument is pinned instead. The docstring says "stream", and this is what
    makes that true.
    """
    path = workbook(tmp_path / "f.xlsx", [["I-200-001", "Certified", None, "A"]])
    seen: list[dict] = []
    original = ingest.load_workbook

    def record(target, **kwargs):
        seen.append(kwargs)
        return original(target, **kwargs)

    monkeypatch.setattr(ingest, "load_workbook", record)
    ingest.read_xlsx(path)
    assert seen and all(k.get("read_only") for k in seen), seen


# --------------------------------------------------------------------------
# The Parquet cache
# --------------------------------------------------------------------------


def test_the_cache_is_written_on_first_read_and_reused_after(source):
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    cache = interim / f"{path.stem}.parquet"

    assert not cache.exists()
    first = ingest.load_raw(path, interim)
    assert cache.is_file()

    # Delete the spreadsheet: a second read that touched it would now fail.
    path.unlink()
    pd.testing.assert_frame_equal(ingest.load_raw(path, interim), first)


def test_a_damaged_cache_is_rebuilt_once_and_the_read_succeeds(source, capsys):
    """The read is the integrity test; a footer check misses damage inside."""
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    ingest.load_raw(path, interim)

    cache = interim / f"{path.stem}.parquet"
    cache.write_bytes(b"not a parquet file")

    assert len(ingest.load_raw(path, interim)) == 2
    assert "rebuilding" in capsys.readouterr().out


def test_a_cache_that_cannot_be_rebuilt_raises_rather_than_looping(
    source, monkeypatch
):
    """Exactly one rebuild. Retrying forever would hang the whole pipeline."""
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    ingest.load_raw(path, interim)
    (interim / f"{path.stem}.parquet").write_bytes(b"still not parquet")

    calls = []
    monkeypatch.setattr(ingest, "_build", lambda *a: calls.append(a))
    with pytest.raises(Exception):
        ingest.load_raw(path, interim)
    assert len(calls) == 1


def test_text_columns_are_cached_as_string_not_object(source):
    """The other half of a bug fixed in clean.py, and the half nobody pinned.

    This cast is what makes the cache hand back ``string`` columns, which is
    why ``clean.to_wage`` exists: ``pd.to_numeric`` on a ``string`` column
    returns a *nullable* dtype whose ``pd.NA`` survives comparisons and turns
    every downstream flag into NA. Without the cast the dtype is ``object``
    and that whole story quietly stops applying.
    """
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    frame = ingest.load_raw(path, interim)
    assert frame["EMPLOYER_NAME"].dtype == "string"
    assert frame["CASE_NUMBER"].dtype == "string"


def test_only_the_requested_columns_come_back(source):
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    frame = ingest.load_raw(path, interim, ["CASE_NUMBER", "EMPLOYER_NAME"])
    assert list(frame.columns) == ["CASE_NUMBER", "EMPLOYER_NAME"]


def test_a_failed_conversion_leaves_no_scratch_file(tmp_path, monkeypatch):
    """A .tmp left in data/ would be committed; interruption here is likely."""
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    path = workbook(raw / "f.xlsx", [["I-200-001", "Certified", "2025-01-15", "A"]])

    monkeypatch.setattr(ingest, "read_xlsx", lambda p: 1 / 0)
    with pytest.raises(ZeroDivisionError):
        ingest.load_raw(path, interim)
    assert list(interim.iterdir()) == []


def test_a_conversion_that_dies_mid_write_leaves_the_old_cache_intact(
    source, monkeypatch
):
    """The failure has to land *during* the write for this to mean anything.

    Raising before the write proves nothing: no destination is touched either
    way. So the fake writes rubbish to whatever path it is given and then
    fails, which is what a half-finished write looks like. Only the rename
    keeps that rubbish away from the real cache.
    """
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    ingest.load_raw(path, interim)
    cache = interim / f"{path.stem}.parquet"
    before = cache.read_bytes()

    def half_written(self, target, **kwargs):
        Path(target).write_bytes(b"half a parquet file")
        raise OSError("no space left on device")

    monkeypatch.setattr(pd.DataFrame, "to_parquet", half_written)
    with pytest.raises(OSError):
        ingest._build(path, cache)

    assert cache.read_bytes() == before
    assert len(pd.read_parquet(cache)) == 2
    # The half-written scratch file is the only trace such a failure can
    # leave, and data/interim is where it would sit until someone noticed.
    assert list(interim.glob("*.tmp")) == []


def test_the_cache_is_never_written_in_place(source, monkeypatch):
    """The destination is only ever reached by rename, never by a write.

    Stated directly rather than inferred from a failure, because a fake that
    raises never reaches the line that would prove it: the exception lands on
    the first write whichever path that write was aimed at.
    """
    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    cache = interim / f"{path.stem}.parquet"
    interim.mkdir(exist_ok=True)

    written: list[Path] = []
    original = pd.DataFrame.to_parquet

    def record(self, target, **kwargs):
        written.append(Path(target))
        return original(self, target, **kwargs)

    monkeypatch.setattr(pd.DataFrame, "to_parquet", record)
    ingest._build(path, cache)

    assert written, "nothing was written at all"
    assert cache not in written, f"{cache.name} was written directly"
    assert all(target.suffix == ".tmp" for target in written), written


def test_conversions_running_at_once_do_not_collide(source):
    """Scratch names carry a uuid, not just the pid: threads share a pid.

    Sequential rebuilds cannot show this — they never overlap. Four threads
    on one destination will, and with a shared scratch name the first rename
    pulls the file out from under the others.
    """
    import threading

    raw, interim = source
    path = next(iter(raw.glob("*.xlsx")))
    interim.mkdir(exist_ok=True)
    cache = interim / f"{path.stem}.parquet"
    errors: list[Exception] = []

    def go():
        try:
            ingest._build(path, cache)
        except Exception as exc:  # noqa: BLE001 - any of them is the finding
            errors.append(exc)

    threads = [threading.Thread(target=go) for _ in range(4)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

    assert errors == []
    assert len(pd.read_parquet(cache)) == 2
    assert list(interim.glob("*.tmp")) == []


# --------------------------------------------------------------------------
# Combining the files
# --------------------------------------------------------------------------


def test_load_all_concatenates_every_source_file(tmp_path):
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    workbook(raw / "a.xlsx", [["I-200-001", "Certified", "2025-01-15", "ACME"]])
    workbook(raw / "b.xlsx", [["I-200-002", "Certified", "2025-04-15", "BETA"]])
    assert sorted(ingest.load_all(raw, interim)["CASE_NUMBER"]) == [
        "I-200-001",
        "I-200-002",
    ]


def test_a_case_in_two_files_keeps_its_later_decision(tmp_path):
    """20,873 real cases span a quarter boundary, always Certified then Withdrawn.

    Ordered by decision date rather than by filename: the two agree today, but
    nothing enforces that, and the duplicates differ in CASE_STATUS.
    """
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    # Deliberately named so alphabetical order disagrees with decision order:
    # sorting by filename alone would keep the *earlier* decision.
    workbook(raw / "a.xlsx", [["I-200-001", "Certified - Withdrawn", "2025-04-02", "A"]])
    workbook(raw / "b.xlsx", [["I-200-001", "Certified", "2025-03-30", "A"]])

    frame = ingest.load_all(raw, interim)
    assert len(frame) == 1
    assert frame["CASE_STATUS"].iloc[0] == "Certified - Withdrawn"


def test_the_columns_needed_for_deduplication_are_always_read(tmp_path):
    """Asking for neither still has to sort by DECISION_DATE to dedupe."""
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    workbook(raw / "a.xlsx", [["I-200-001", "Certified", "2025-01-15", "ACME"]])
    frame = ingest.load_all(raw, interim, ["EMPLOYER_NAME"])
    assert {"CASE_NUMBER", "DECISION_DATE"} <= set(frame.columns)


def test_the_index_is_reset_after_deduplication(tmp_path):
    """A gapped index turns later positional work into silent misalignment.

    Rows are written out of order and one is duplicated, so the sort and the
    deduplication both move things. Already-sorted input would come back with
    a tidy index whether or not it was reset.
    """
    raw, interim = tmp_path / "raw", tmp_path / "interim"
    raw.mkdir()
    workbook(
        raw / "a.xlsx",
        [
            ["I-200-009", "Certified", "2025-01-19", "ACME"],
            ["I-200-003", "Certified", "2025-01-13", "BETA"],
            ["I-200-009", "Certified - Withdrawn", "2025-02-19", "ACME"],
            ["I-200-001", "Certified", "2025-01-11", "GAMMA"],
        ],
    )
    frame = ingest.load_all(raw, interim)
    assert len(frame) == 3
    assert list(frame.index) == list(range(len(frame)))
