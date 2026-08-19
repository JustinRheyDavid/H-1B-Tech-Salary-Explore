"""Read raw DOL LCA disclosure files.

The source files are large, slow to parse, and inconsistent in ways that are
documented in ``notebooks/01_exploration.ipynb``. This module hides three of
those problems:

* sheet names differ in every file, so the sheet is chosen by index
* roughly 73% of rows are blank padding, so they are dropped on read
* DOL misspelled one filename, so sources are matched by glob

Reading all nine files takes about 15 minutes. Each one is converted to
Parquet once and reused, which brings a reload down to seconds.
"""

from __future__ import annotations

import os
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

__all__ = ["read_xlsx", "load_raw", "load_all", "source_files"]


def source_files(raw_dir: Path) -> list[Path]:
    """Every disclosure spreadsheet in ``raw_dir``, in a stable order.

    Matched by glob rather than by name: the FY2026 file is published as
    ``LCA_Dislclosure_Data_FY2026_Q2.xlsx``, misspelling "Disclosure".
    """
    files = sorted(Path(raw_dir).glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(
            f"No .xlsx files found in {raw_dir}. "
            "See the Data sources section of README.md for what to download."
        )
    return files


def read_xlsx(path: Path) -> pd.DataFrame:
    """Stream one spreadsheet into a DataFrame, dropping blank padding rows.

    The sheet is taken by index because names are inconsistent across files
    (``Q1``, ``LCA_Disclosure_Data_FY2025_Q1``, and ``Sheet1`` all occur).
    """
    path = Path(path)
    workbook = load_workbook(path, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        if "CASE_NUMBER" not in header:
            raise ValueError(
                f"{path.name}: no CASE_NUMBER column; got {header[:5]}..."
            )
        key = header.index("CASE_NUMBER")
        data = [row for row in rows if row[key] is not None]
    finally:
        workbook.close()
    return pd.DataFrame(data, columns=header)


def _build(path: Path, dest: Path) -> Path:
    """Convert ``path`` to Parquet at ``dest``, replacing it atomically.

    The scratch name is unique per call. A shared scratch path lets the first
    rename pull the file out from under a concurrent writer, and the process
    id alone is not enough because threads in one process share it.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(f".parquet.{os.getpid()}.{uuid4().hex[:8]}.tmp")
    try:
        frame = read_xlsx(path)
        frame = frame.astype(
            {c: "string" for c in frame.columns if frame[c].dtype == object}
        )
        frame.to_parquet(tmp, compression="snappy", index=False)
        tmp.replace(dest)  # overwrites in place, so a locked dest cannot strand us
    finally:
        tmp.unlink(missing_ok=True)  # tidy up if the conversion raised
    return dest


def load_raw(
    path: Path,
    interim_dir: Path,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Read one source file through its Parquet cache.

    The read is the integrity test. A footer check would catch a truncated
    file but not damage inside a row group, so a failed read triggers exactly
    one rebuild before giving up.
    """
    path = Path(path)
    dest = Path(interim_dir) / (path.stem + ".parquet")
    if not dest.exists():
        _build(path, dest)
    try:
        return pd.read_parquet(dest, columns=columns)
    except Exception as exc:  # corrupt cache: rebuild once, then let it raise
        print(f"  cache for {path.name} unreadable ({type(exc).__name__}); rebuilding")
        _build(path, dest)
        return pd.read_parquet(dest, columns=columns)


def load_all(
    raw_dir: Path,
    interim_dir: Path,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Read every source file, concatenate, and deduplicate on case number.

    20,873 cases appear in two files, always spanning a quarter boundary and
    always as a ``Certified`` to ``Certified - Withdrawn`` transition. Sorting
    by ``DECISION_DATE`` makes "keep the later state" explicit rather than an
    accident of how the files happen to be named.
    """
    frames = [
        load_raw(p, interim_dir, needed_columns(columns))
        for p in source_files(raw_dir)
    ]
    return combine(frames)


def needed_columns(columns: list[str] | None) -> list[str] | None:
    """``columns`` plus the two the deduplication itself needs.

    Reading a narrower set than :func:`combine` requires fails inside pandas
    with a ``KeyError`` on a column the caller never asked about.
    """
    if columns is None:
        return None
    return sorted(set(columns) | {"CASE_NUMBER", "DECISION_DATE"})


def combine(frames: list[pd.DataFrame]) -> pd.DataFrame:
    """Concatenate source frames and resolve cases that appear in two of them.

    Split out from :func:`load_all` so that a caller holding the Parquet caches
    *without* the spreadsheets they were built from — the Azure ETL container,
    which downloads the caches from Blob and never sees an ``.xlsx`` — gets this
    exact logic rather than a second implementation of it.

    That matters more than it looks: 20,873 cases appear twice, so a loader that
    concatenated without this step would load 871,194 rows and every count in
    the runbook would be wrong by the same amount.
    """
    combined = pd.concat(frames, ignore_index=True)
    return (
        combined.sort_values(["CASE_NUMBER", "DECISION_DATE"])
        .drop_duplicates("CASE_NUMBER", keep="last")
        .reset_index(drop=True)
    )
